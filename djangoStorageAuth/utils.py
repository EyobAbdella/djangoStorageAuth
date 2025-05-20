from django.conf import settings
from django.utils import timezone
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from .models import OAuthTokens
import os
import msal
import requests
import re

def load_google_credentials(user_id):
    oauth_tokens = OAuthTokens.objects.filter(user_id=user_id).only("google_access", "google_refresh").first()
    if not oauth_tokens:
        return None
    credentials = Credentials(
        token=oauth_tokens.google_access,
        refresh_token=oauth_tokens.google_refresh,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=settings.GOOGLE_CLIENT_ID,
        client_secret=settings.GOOGLE_CLIENT_SECRET,
        scopes=[
            "https://www.googleapis.com/auth/drive",
            "https://www.googleapis.com/auth/spreadsheets"
        ],
    )

    if credentials.expired and credentials.refresh_token:
        credentials.refresh(Request())
        OAuthTokens.objects.filter(user_id=user_id).update(google_access=credentials.token)

    return credentials


def list_drive_files(user_id, mime_type=None):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return []

    drive_service = build("drive", "v3", credentials=credentials)

    query = "trashed=false"
    if mime_type:
        query += f" and mimeType='{mime_type}'"

    response = drive_service.files().list(
        q=query,
        spaces='drive',
        fields="files(id, name, mimeType)",
        pageSize=100,
    ).execute()

    return response.get("files", [])

def list_google_drive_folders(user_id):
    return list_drive_files(user_id, mime_type="application/vnd.google-apps.folder")

def create_drive_file(user_id, title="Untitled File", mime_type="application/vnd.google-apps.document", parent_folder_id=None):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    drive_service = build("drive", "v3", credentials=credentials)

    file_metadata = {
        "name": title,
        "mimeType": mime_type
    }

    if parent_folder_id:
        file_metadata["parents"] = [parent_folder_id]

    file = drive_service.files().create(
        body=file_metadata,
        fields="id, name, mimeType, webViewLink"
    ).execute()

    return {
        "id": file["id"],
        "name": file["name"],
        "mime_type": file["mimeType"],
        "url": file["webViewLink"]
    }

def update_drive_file_metadata(user_id, file_id, new_title=None, new_parent_folder_id=None):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    drive_service = build("drive", "v3", credentials=credentials)

    updates = {}
    if new_title:
        updates["name"] = new_title

    if updates:
        drive_service.files().update(
            fileId=file_id,
            body=updates,
            fields="id, name, mimeType"
        ).execute()

    if new_parent_folder_id:
        file = drive_service.files().get(fileId=file_id, fields='parents').execute()
        previous_parents = ",".join(file.get('parents', []))
        
        drive_service.files().update(
            fileId=file_id,
            addParents=new_parent_folder_id,
            removeParents=previous_parents,
            fields="id, parents"
        ).execute()

    return {"status": "success", "file_id": file_id}

def delete_drive_file(user_id, file_id):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    drive_service = build("drive", "v3", credentials=credentials)

    try:
        drive_service.files().delete(fileId=file_id).execute()
        return {"status": "success", "file_id": file_id}
    except:
        return None

def add_data_to_sheet(user_id, sheet_id, range_name, values):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    sheets_service = build("sheets", "v4", credentials=credentials)

    try:
        body = {
            'values': values
        }
        result = sheets_service.spreadsheets().values().append(
            spreadsheetId=sheet_id,
            range=range_name,
            valueInputOption="RAW",
            body=body
        ).execute()
        return {
            "status": "success",
            "sheet_id": sheet_id,
            "updated_range": result.get("updates", {}).get("updatedRange")
        }
    except:
        return None

def delete_data_from_sheet(user_id, sheet_id, operation, range_name=None, row_indices=None, sheet_name="Sheet1"):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    sheets_service = build("sheets", "v4", credentials=credentials)

    try:
        if operation == "bulk":
            sheets_service.spreadsheets().values().clear(
                spreadsheetId=sheet_id,
                range=f"{sheet_name}!A1:Z1000"
            ).execute()
            return {"status": "success", "sheet_id": sheet_id, "operation": "bulk_clear"}

        elif operation == "cell" and range_name:
            sheets_service.spreadsheets().values().clear(
                spreadsheetId=sheet_id,
                range=range_name
            ).execute()
            return {"status": "success", "sheet_id": sheet_id, "operation": "cell_clear", "range": range_name}

        elif operation in ["row", "rows"] and row_indices:
            row_indices = sorted(row_indices, reverse=True)
            requests = [
                {
                    "deleteDimension": {
                        "range": {
                            "sheetId": 0,
                            "dimension": "ROWS",
                            "startIndex": index - 1,
                            "endIndex": index
                        }
                    }
                } for index in row_indices
            ]
            body = {"requests": requests}
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=sheet_id,
                body=body
            ).execute()
            return {
                "status": "success",
                "sheet_id": sheet_id,
                "operation": "row_delete",
                "deleted_rows": row_indices
            }

        return None
    except:
        return None

def update_data_in_sheet(user_id, sheet_id, range_name, values):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    sheets_service = build("sheets", "v4", credentials=credentials)

    try:
        body = {
            'values': values
        }
        result = sheets_service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=range_name,
            valueInputOption="RAW",
            body=body
        ).execute()
        return {
            "status": "success",
            "sheet_id": sheet_id,
            "updated_range": result.get("updatedRange"),
            "updated_rows": result.get("updatedRows"),
            "updated_columns": result.get("updatedColumns")
        }
    except:
        return None

def list_data_from_sheet(user_id, sheet_id, range_name="Sheet1!A1:Z1000", search_column=None, search_keyword=None):
    credentials = load_google_credentials(user_id)
    if not credentials:
        return None

    sheets_service = build("sheets", "v4", credentials=credentials)

    try:
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=range_name
        ).execute()
        values = result.get("values", [])

        if not values:
            return {"status": "success", "sheet_id": sheet_id, "data": []}

        if search_column is not None and search_keyword is not None:
            try:
                search_column_index = ord(search_column.upper()) - ord('A')
            except:
                return None

            filtered_values = [
                row for row in values
                if len(row) > search_column_index and str(row[search_column_index]).lower() == search_keyword.lower()
            ]
            return {
                "status": "success",
                "sheet_id": sheet_id,
                "data": filtered_values,
                "search_column": search_column,
                "search_keyword": search_keyword
            }

        return {"status": "success", "sheet_id": sheet_id, "data": values}
    except:
        return None

# Microsoft Functions
def load_microsoft_credentials(user_id):
    oauth_tokens = OAuthTokens.objects.filter(user_id=user_id).only("microsoft_access", "microsoft_refresh", "microsoft_expiry").first()
    if not oauth_tokens:
        return None

    app = msal.ConfidentialClientApplication(
        client_id=settings.MICROSOFT_CLIENT_ID,
        client_credential=settings.MICROSOFT_CLIENT_SECRET,
        authority="https://login.microsoftonline.com/common"
    )

    # check if token is expired (with 5-minute buffer)
    if oauth_tokens.microsoft_expiry <= timezone.now() + timezone.timedelta(minutes=5) and oauth_tokens.microsoft_refresh:
        result = app.acquire_token_by_refresh_token(
            refresh_token=oauth_tokens.microsoft_refresh,
            scopes=["User.Read", "Files.ReadWrite.All"]
        )
        if "access_token" in result:
            oauth_tokens.microsoft_access = result["access_token"]
            oauth_tokens.microsoft_refresh = result.get("refresh_token", oauth_tokens.microsoft_refresh)
            oauth_tokens.microsoft_expiry = timezone.now() + timezone.timedelta(seconds=result.get("expires_in", 3600))
            oauth_tokens.save()
    elif not oauth_tokens.microsoft_access:
        return None

    return oauth_tokens.microsoft_access

def list_onedrive_files(user_id, mime_type=None):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return []

    try:
        query_params = {
            '$select': "id,name,file",
            '$top': 100
        }
        if mime_type:
            query_params['$filter'] = f"file/mimeType eq '{mime_type}'"

        response = requests.get(
            "https://graph.microsoft.com/v1.0/me/drive/root/children",
            params=query_params,
            headers={"Authorization": f"Bearer {access_token}"}
        )
        if response.status_code != 200:
            return []
        files = []
        for item in response.json().get("value", []):
            if item.get("file") and not item.get("folder"):
                files.append({
                    "id": item["id"],
                    "name": item["name"],
                    "mime_type": item["file"]["mimeType"]
                })
        return files
    except Exception as e:
        return []

def list_onedrive_folders(user_id):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return []

    try:
        response = requests.get(
            "https://graph.microsoft.com/v1.0/me/drive/root/children",
            params={
                'filter': "folder ne null",
                'select': "id,name"
            },
            headers={"Authorization": f"Bearer {access_token}"}
        )
        if response.status_code != 200:
            return []
        return [
            {"id": item["id"], "name": item["name"], "mime_type": "application/vnd.google-apps.folder"}
            for item in response.json().get("value", [])
        ]
    except Exception as e:
        return []

def create_onedrive_file(user_id, title="Untitled File", mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", parent_folder_id=None):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            template_path = os.path.join(settings.BASE_DIR, "templates", "empty_excel_template.xlsx")
            if not os.path.exists(template_path):
                wb = Workbook()
                wb.save(template_path)

            with open(template_path, "rb") as f:
                path = f"/me/drive/items/{parent_folder_id}:/{title}.xlsx:/content" if parent_folder_id else f"/me/drive/root:/{title}.xlsx:/content"
                response = requests.put(
                    f"https://graph.microsoft.com/v1.0{path}",
                    data=f,
                    headers={
                        "Content-Type": mime_type,
                        "Authorization": f"Bearer {access_token}"
                    }
                )
        else:
            metadata = {
                "name": title,
                "file": {},
                "@microsoft.graph.conflictBehavior": "rename"
            }
            path = f"/me/drive/items/{parent_folder_id}/children" if parent_folder_id else "/me/drive/root/children"
            response = requests.post(
                f"https://graph.microsoft.com/v1.0{path}",
                json=metadata,
                headers={"Authorization": f"Bearer {access_token}"}
            )

        if response.status_code not in [200, 201]:
            return None
        file_info = response.json()
        file_id = file_info["id"]
        return {
            "id": file_id,
            "name": file_info["name"],
            "mime_type": file_info.get("file", {}).get("mimeType", mime_type),
            "url": file_info.get("webUrl")
        }
    except Exception as e:
        return None
def update_onedrive_file_metadata(user_id, file_id, new_title=None, new_parent_folder_id=None):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    if not file_id:
        return None

    if not new_title and not new_parent_folder_id:
        return None

    try:
        if new_title:
            response = requests.patch(
                f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}",
                json={"name": new_title},
                headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
            )
            if response.status_code != 200:
                return None

        if new_parent_folder_id:
            response = requests.patch(
                f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}",
                json={"parentReference": {"id": new_parent_folder_id}},
                headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
            )
            if response.status_code != 200:
                return None

        return {"status": "success", "file_id": file_id}
    except Exception as e:
        return None
def delete_onedrive_file(user_id, file_id):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        response = requests.delete(
            f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}",
            headers={"Authorization": f"Bearer {access_token}"}
        )
        if response.status_code != 204:
            return None
        return {"status": "success", "file_id": file_id}
    except Exception as e:
        return None

def add_data_to_excel(user_id, workbook_id, range_name, values, sheet_name="Sheet1"):
    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        list_sheets_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets"
        response = requests.get(list_sheets_url, headers=headers)
        if response.status_code != 200:
            return None

        existing_sheets = [sheet["name"] for sheet in response.json().get("value", [])]
        if sheet_name not in existing_sheets:
            create_sheet_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets/add"
            sheet_body = {"name": sheet_name}
            response = requests.post(create_sheet_url, headers=headers, json=sheet_body)
            if response.status_code not in [200, 201]:
                return None

        num_rows = len(values)
        num_cols = len(values[0]) if values else 0
        _, cell_ref = range_name.split("!")
        start_col_letter = cell_ref[0].upper()
        start_row = int(cell_ref[1:])
        end_col_letter = chr(ord(start_col_letter) + num_cols - 1) if num_cols > 0 else start_col_letter
        end_row = start_row + num_rows - 1
        range_address = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"

        graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='{range_address}')"
        formatted_data = {"values": values}
        response = requests.patch(graph_url, headers=headers, json=formatted_data)
        if response.status_code != 200:
            return None

        return {
            "status": "success",
            "workbook_id": workbook_id,
            "updated_range": range_name
        }
    except Exception as e:
        return None

def update_data_in_excel(user_id, workbook_id, range_name, values):

    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        if not re.match(r"^[A-Za-z0-9]+![A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", range_name):
            return None

        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        try:
            sheet_name, cell_ref = range_name.split("!")
        except ValueError:
            return None

        list_sheets_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets"
        response = requests.get(list_sheets_url, headers=headers)
        if response.status_code != 200:
            return None

        existing_sheets = [sheet["name"] for sheet in response.json().get("value", [])]
        if sheet_name not in existing_sheets:
            return None

        num_rows = len(values)
        num_cols = len(values[0]) if values else 0
        try:
            start_col_letter = cell_ref[0].upper()
            start_row = int(cell_ref[1:])
            end_col_letter = chr(ord(start_col_letter) + num_cols - 1) if num_cols > 0 else start_col_letter
            end_row = start_row + num_rows - 1
            range_address = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        except (ValueError, IndexError) as e:
            return None

        graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='{range_address}')"
        formatted_data = {"values": values}
        response = requests.patch(graph_url, headers=headers, json=formatted_data)
        if response.status_code != 200:
            return None

        return {
            "status": "success",
            "workbook_id": workbook_id,
            "updated_range": range_name,
            "updated_rows": num_rows,
            "updated_columns": num_cols
        }
    except Exception as e:
        return None

def delete_data_from_excel(user_id, workbook_id, operation, range_name=None, row_indices=None, sheet_name="Sheet1"):

    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        list_sheets_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets"
        response = requests.get(list_sheets_url, headers=headers)
        if response.status_code != 200:
            return None

        existing_sheets = [sheet["name"] for sheet in response.json().get("value", [])]
        if sheet_name not in existing_sheets:
            return None

        if operation == "bulk":
            range_address = "A1:Z1000"
            graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='{range_address}')/clear"
            response = requests.post(graph_url, headers=headers, json={"applyTo": "contents"})
            if response.status_code not in [200, 204]:
                return None
            result = {"status": "success", "workbook_id": workbook_id, "operation": "bulk_clear"}

        elif operation == "cell" and range_name:
            try:
                _, cell_ref = range_name.split("!")
            except ValueError:
                return None
            graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='{cell_ref}')/clear"
            response = requests.post(graph_url, headers=headers, json={"applyTo": "contents"})
            if response.status_code not in [200, 204]:
                return None
            result = {"status": "success", "workbook_id": workbook_id, "operation": "cell_clear", "range": range_name}

        elif operation in ["row", "rows"] and row_indices:
            row_indices = sorted(row_indices, reverse=True)
            for index in row_indices:
                graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='A{index}:Z{index}')/delete"
                response = requests.post(graph_url, headers=headers, json={"shift": "Up"})
                if response.status_code not in [200, 204]:
                    return None
            result = {
                "status": "success",
                "workbook_id": workbook_id,
                "operation": "row_delete",
                "deleted_rows": row_indices
            }

        else:
            return None

        return result
    except Exception as e:
        return None
def list_data_from_excel(user_id, workbook_id, range_name="Sheet1!A1:Z100", search_column=None, search_keyword=None):

    access_token = load_microsoft_credentials(user_id)
    if not access_token:
        return None

    try:
        if not re.match(r"^[A-Za-z0-9]+![A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", range_name):
            return None

        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        try:
            sheet_name, _ = range_name.split("!")
        except ValueError:
            return None

        list_sheets_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets"
        response = requests.get(list_sheets_url, headers=headers)
        if response.status_code != 200:
            return None

        existing_sheets = [sheet["name"] for sheet in response.json().get("value", [])]
        if sheet_name not in existing_sheets:
            return None

        graph_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{workbook_id}/workbook/worksheets('{sheet_name}')/range(address='{range_name}')"
        response = requests.get(graph_url, headers=headers)
        if response.status_code != 200:
            return None

        values = response.json().get("values", [])
        values = [
            [cell for cell in row if cell is not None and str(cell).strip()]
            for row in values
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        if not values:
            return {"status": "success", "workbook_id": workbook_id, "data": []}

        if search_column is not None and search_keyword is not None:
            try:
                search_column_index = ord(search_column.upper()) - ord('A')
                filtered_values = [
                    row for row in values
                    if len(row) > search_column_index and str(row[search_column_index]).lower() == search_keyword.lower()
                ]
                return {
                    "status": "success",
                    "workbook_id": workbook_id,
                    "data": filtered_values,
                    "search_column": search_column,
                    "search_keyword": search_keyword
                }
            except Exception as e:
                return None

        return {"status": "success", "workbook_id": workbook_id, "data": values}
    except Exception as e:
        return None